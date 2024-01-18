import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl import Workbook, load_workbook

wb_test = Workbook()
wb_test.save('carx.xlsx')
ws = wb_test.active

ws = wb_test.create_sheet('Sheet1')

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

auto = str(input("Ievadiet automobila marku: "))
modelis = str(input("Ievadiet automobila modeli: "))

url = "https://www.ss.com/lv/transport/cars/"
driver.get(url)
time.sleep(2)

find = driver.find_element(By.LINK_TEXT, auto)
find.click()
time.sleep(2)

find = driver.find_element(By.LINK_TEXT, "Meklēšana")
find.click()
time.sleep(2)

find = driver.find_element(By.ID, "ptxt")
find.clear()
find.send_keys(modelis)
time.sleep(3)
find.send_keys(Keys.RETURN)

cena_list = []
tilp_list = []
nobr_list = []
gad_list = []
mod_list = []

kollonas = {
    'cena' : 4,
    'tilp' : 3,
    'nobr' : 2,
    'gad' : 1,
    'mod' : 0,
}

for column, offset in kollonas.items():
    i = 0
    while i <=34:
        x = driver.find_elements(By.CLASS_NAME, 'amopt')

        if offset >= len(x):
            break

        element = x[offset]
        text_vert = element.text

        if column == 'cena':
            text_vert = text_vert.replace('€', '').replace(',','.').replace('/n','').replace('maiņai','').replace('maiņa','').replace('pērku','')
            vert = float(text_vert)
        else:
            vert = text_vert
        globals()[f'{column}_list'].append(vert)

        offset += 5
        i += 1

print(cena_list)
print(nobr_list)
print(tilp_list)
print(gad_list)
print(mod_list)

valid_price = [price for price in cena_list if price is not None]
average_price = sum(valid_price) / len(valid_price) if valid_price else 0

print('\nPiedavajumi zem videjas tirgus cenas: ')
for i, price in enumerate(cena_list):
    if price is not None and price < average_price:
        print(f"{nobr_list[i]}, {mod_list[i]}, {gad_list[i]}, {tilp_list[i]}, {price} tūkst.€")
        ws.append((nobr_list[i], mod_list[i], gad_list[i], tilp_list[i], cena_list[i]))
        
wb_test.save('carx.xlsx')
wb_test.close

driver.quit()
input()