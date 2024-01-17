
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import Workbook, load_workbook 



service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

name=[]
result = []

with open("people.csv", "r") as file:
    next(file)
    for line in file:
        row=line.rstrip().split(",") 
        full_name = ' '.join((row[2].replace(",", " "), row[3].replace(",", " ")))
        name.append(full_name)
        print(full_name)

url = "https://emn178.github.io/online-tools/crc32.html"
driver.get(url)
time.sleep(2)


for names in name:
    find = driver.find_element(By.ID, "input")
    find.clear()
    find.send_keys(names)
    find = driver.find_element(By.ID, "output")
    x = find.get_attribute("value")
    print(x)
    result.append(x)

print(result)

input()

fails = 'salary.xlsx'
darbinieki = {}

wb = openpyxl.load_workbook(fails)
sheet = wb.active


for row in sheet.iter_rows(min_row=2, values_only=True):

    vards_kodets, alga = row

    vards = None
    for darbinieka_vards, kodets in zip(name, result):
        if kodets == vards_kodets:
            vards = darbinieka_vards
            break

    if vards is not None:

        if vards in darbinieki:
            darbinieki[vards] += alga
        else:
            darbinieki[vards] = alga

for vards, alga_sum in darbinieki.items():
    print(f"{vards}: {alga_sum}")


